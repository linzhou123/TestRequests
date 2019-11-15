import openpyxl
import os
from openpyxl import load_workbook,Workbook

class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename,sheetname):
        self.filename = filename
        self.sheetname=sheetname
        #判断是否存在这个文件，存在读取写入，不存在生成文件
        if not os.path.exists(self.filename):
            self.wb=Workbook()
        else:
            self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet
    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)
